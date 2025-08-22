# import rawpy
# import sys
# import os
# import cv2
# import matplotlib
# matplotlib.use('TkAgg')  # 或 'Agg' (非互動), 'WXAgg'，等等
# import matplotlib.pyplot as plt

# sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
# from common.common_func import show_image, show_two_images, show_three_images, show_four_images



# raw_short_path = "C:/Users/eevo1/OneDrive/Desktop/ISP_AI_Comparison/data/raw/Fuji/Fuji/short/00001_00_0.1s.RAF"
# with rawpy.imread(raw_short_path) as raw:
#     small_bayer = cv2.resize(raw.raw_image_visible, (640, 480))
    

# raw_long_path = "C:/Users/eevo1/OneDrive/Desktop/ISP_AI_Comparison/data/raw/Fuji/Fuji/long/00001_00_10s.RAF"
# with rawpy.imread(raw_long_path) as raw:
#     rgb_long = raw.postprocess(
#         use_camera_wb=True,      # 使用相機內建白平衡
#         no_auto_bright=True,     # 不自動調整亮度
#         output_bps=8             # 輸出8位元影像
#     )


# show_two_images(
#     small_bayer, 'Bayer Raw Image',     
#     rgb_long, 'Long Exposure RGB Image',
#     cmap1=None, cmap2=None
# )


import openpyxl
from openpyxl.styles import Font, PatternFill

# 建立工作簿與工作表
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ISP Project Plan"

# 標題行
headers = ["週數", "工作項目", "詳細說明", "使用語言/工具", "目標成果", "備註"]
ws.append(headers)

# 調整標題欄位寬度
col_widths = [8, 30, 60, 20, 30, 30]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# 文字置中且換行
alignment_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

for col in range(1, len(headers)+1):
    ws.cell(row=1, column=col).alignment = alignment_wrap

# 計劃內容，依週數拆細
plan_data = [
    # 週, 工作項目, 詳細說明, 語言/工具, 目標成果, 備註
    [1, "環境建置", "安裝Python、OpenCV、rawpy、matplotlib等必要套件", "Python", "可順利執行示範程式", "使用conda環境管理"],
    [1, "資料準備", "準備不同曝光時間的RAW檔案，確保讀取無誤", "Python, rawpy", "RAW圖成功讀取並顯示", "測試rawpy.postprocess"],
    
    [2, "基本讀取與顯示", "實作RAW檔案讀取及bayer影像顯示", "Python, matplotlib", "顯示bayer raw及rawpy postprocess結果", "確認顏色與形狀正常"],
    [2, "基本去馬賽克", "使用OpenCV去馬賽克(Bayer轉RGB)", "Python, OpenCV", "得到去馬賽克彩色影像", "結果與rawpy.postprocess比較"],
    
    [3, "Gamma校正", "實作並套用Gamma校正函數，改善影像對比", "Python, numpy", "改善後影像對比度提升", "定義class封裝函數"],
    [3, "正規化與色彩調整", "實作影像正規化至0-255與uint8轉換", "Python, numpy", "影像能正確顯示色彩", "與Gamma校正搭配使用"],
    
    [4, "自製傳統ISP - 去噪", "加入簡易去噪模組（如中值濾波）", "Python, OpenCV", "影像噪聲降低，細節保持", "參考影像品質改進"],
    [4, "自製傳統ISP - 白平衡", "實作白平衡演算法", "Python, numpy", "影像色彩更自然", "可用灰世界法或其他簡單方法"],
    
    [5, "自製傳統ISP - 色彩校正", "加入色彩校正矩陣，模擬相機ISP效果", "Python, numpy", "色彩準確度提升", "參考相機白平衡結果"],
    [5, "效果比較", "整理bayer、rawpy postprocess、OpenCV去馬賽克、自製ISP結果", "Python, matplotlib", "生成多圖比較視覺報告", "便於分析與調整"],
    
    [6, "AI輔助ISP - 讀取模型", "簡易整合預訓練AI模型輸入輸出", "Python, PyTorch/TensorFlow", "AI處理後影像生成", "與傳統方法比較"],
    [6, "AI輔助ISP - 效果優化", "嘗試不同AI模型參數調整", "Python", "提升AI影像品質", "可微調提升細節與對比"],
    
    [7, "離線Demo實作", "製作離線讀圖介面及多影像切換", "Python, tkinter/其他GUI", "可在筆電上展示多種影像處理結果", "無需即時影像串流"],
    [7, "基本佈署準備", "測試程式在筆電Ubuntu環境跑通", "Python, Ubuntu", "能於Linux環境順利運行", "準備交叉編譯基礎"],
    
    [8, "交叉編譯測試", "使用交叉編譯工具將程式編譯至目標嵌入式Linux", "C/C++或Python", "能在目標板運行基本功能", "準備面試Demo材料"],
    [8, "Demo腳本完善", "完善演示腳本與使用說明", "Python", "簡單易懂的Demo流程", "含說明文件"],
]

# 寫入資料並設定自動換行
for row in plan_data:
    ws.append(row)
    current_row = ws.max_row
    for col in range(1, len(headers)+1):
        ws.cell(row=current_row, column=col).alignment = alignment_wrap

# 儲存檔案
output_path = "/mnt/data/ISP_Project_Plan_Weekly.xlsx"
wb.save(output_path)
print(f"Excel檔案已生成並儲存在: {output_path}")
